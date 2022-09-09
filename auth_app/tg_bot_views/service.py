import datetime
import logging
from typing import Dict, Tuple

from django.contrib.auth import get_user_model
from django.db.models import Count, Q, Sum
from django.db.models.query import QuerySet
from openpyxl import Workbook
from openpyxl.styles import Alignment

from auth_app.models import Period, Transaction, UserStat

User = get_user_model()

logger = logging.getLogger(__name__)

ADMIN_REPORT_DATA_CELLS = "BCDEFGHIJK"
ADMIN_REPORT_DATA_HEADERS = (
    'Доступно', 'Получено', 'Отправлено',
    'Сгорят', 'Проверка баланса', 'Входящие (ожидание)',
    'Исходящие (ожидание)', 'Начальный баланс',
    'Начислений', 'Распределений'
)

USER_REPORT_DATA_CELLS = "ABCDEFG"
USER_REPORT_DATA_HEADERS = (
    'Создана', 'ID получателя', 'Кому', 'Количество',
    'Статус', 'Анонимность', 'Обоснование'
)


def get_period() -> Period:
    today = datetime.date.today()
    current_period = Period.objects.filter(
        Q(start_date__lte=today) & Q(end_date__gte=today)).first()
    if current_period is None:
        previous_period = (Period.objects.filter(end_date__lt=today)
                           .order_by('-end_date').first())
        return previous_period
    return current_period


def export_user_transactions(telegram_id: str) -> str:
    wb = Workbook()
    ws = wb.active
    period = get_period()
    transactions = get_user_transactions(period, telegram_id)
    make_table_structure_for_user_report(ws)
    for index, transaction in enumerate(transactions, 3):
        ws[f'A{index}'].value = (transaction.created_at + datetime.timedelta(hours=3)).strftime("%Y-%m-%d %H:%M:%S.%f")
        ws[f'B{index}'].value = transaction.recipient.profile.tg_id
        ws[f'C{index}'].value = get_recipient_name_or_return_me(transaction.recipient, telegram_id)
        ws[f'D{index}'].value = transaction.amount
        ws[f'E{index}'].value = transaction.get_status_display()
        ws[f'F{index}'].value = get_yes_or_no_about_anonymous_transaction_status(transaction.is_anonymous)
        ws[f'G{index}'].value = transaction.reason
    filename = f'{get_export_report_filename(telegram_id)}.xlsx'
    wb.save(filename)
    return filename


def get_export_report_filename(telegram_id):
    return f"{telegram_id}_{datetime.datetime.now()}"


def get_yes_or_no_about_anonymous_transaction_status(anonymous_status: bool) -> str:
    if anonymous_status is True:
        return 'Да'
    return 'Нет'


def get_recipient_name_or_return_me(recipient: User, tg_id: str) -> str:
    if recipient.profile.tg_id == tg_id:
        return 'Мне'
    return (f"{recipient.profile.get_surname} "
            f"{recipient.profile.get_first_name} "
            f"{recipient.profile.get_middle_name}")


def get_user_transactions(period: Period, telegram_id: str) -> QuerySet[Transaction]:
    transactions = (Transaction.objects
                    .select_related('sender__profile', 'recipient__profile')
                    .filter((Q(sender__profile__tg_id=telegram_id) |
                             Q(recipient__profile__tg_id=telegram_id)), period=period)).only(
        'created_at', 'amount', 'status', 'is_anonymous', 'reason', 'sender__id',
        'recipient__profile__first_name',
        'recipient__profile__surname',
        'recipient__profile__middle_name',
        'recipient__profile__tg_id'
    )
    return transactions


def make_table_structure_for_user_report(ws: Workbook.active) -> None:
    for i in range(len(USER_REPORT_DATA_CELLS)):
        ws.column_dimensions[USER_REPORT_DATA_CELLS[i]].width = 16
        ws.merge_cells(f"{USER_REPORT_DATA_CELLS[i]}1:{USER_REPORT_DATA_CELLS[i]}2")
        ws[f"{USER_REPORT_DATA_CELLS[i]}1"].value = USER_REPORT_DATA_HEADERS[i]
        ws[f"{USER_REPORT_DATA_CELLS[i]}1"].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)


def create_admin_report() -> str:
    wb = Workbook()
    ws = wb.active
    period = get_period()
    (recipient_counter_data, recipient_waiting_data,
     sender_counter_data, sender_waiting_data, stats) = get_transactions_data(period)
    make_table_structure_for_admin_report(ws)
    for index, stat in enumerate(stats, 6):
        ws[f'A{index}'].value = (f"{stat.user.profile.get_first_name} "
                                 f"{stat.user.profile.get_middle_name} "
                                 f"{stat.user.profile.get_surname}")
        ws[f'B{index}'].value = stat.distr_initial - stat.distr_thanks
        ws[f'C{index}'].value = stat.income_thanks
        ws[f'D{index}'].value = stat.distr_thanks
        ws[f'E{index}'].value = stat.distr_initial - stat.distr_thanks + stat.income_thanks
        ws[f'F{index}'].value = stat.distr_initial - stat.distr_thanks
        ws[f'G{index}'].value = recipient_waiting_data.get(stat.user_id, 0)
        ws[f'H{index}'].value = sender_waiting_data.get(stat.user_id, 0)
        ws[f'I{index}'].value = stat.distr_initial
        ws[f'J{index}'].value = recipient_counter_data.get(stat.user_id, 0)
        ws[f'K{index}'].value = sender_counter_data.get(stat.user_id, 0)
    filename = f"{get_admin_report_filename()}.xlsx"
    wb.save(filename)
    return filename


def get_admin_report_filename():
    return f"report_{datetime.datetime.now()}"


def make_table_structure_for_admin_report(ws: Workbook.active) -> None:
    ws.column_dimensions['A'].width = 28
    for i in range(len(ADMIN_REPORT_DATA_CELLS)):
        ws.column_dimensions[ADMIN_REPORT_DATA_CELLS[i]].width = 14
    ws.merge_cells('A2:A5')
    for i in range(len(ADMIN_REPORT_DATA_CELLS)):
        ws.merge_cells(f'{ADMIN_REPORT_DATA_CELLS[i]}3:{ADMIN_REPORT_DATA_CELLS[i]}5')
    ws.merge_cells('B2:I2')
    ws.merge_cells('J2:K2')
    for i in range(len(ADMIN_REPORT_DATA_CELLS)):
        ws[f'{ADMIN_REPORT_DATA_CELLS[i]}3'].value = ADMIN_REPORT_DATA_HEADERS[i]
    surname_name_cell = ws.cell(row=2, column=1)
    thanks_amount_cell = ws['B2']
    transactions_amount_cell = ws['J2']
    surname_name_cell.value = 'Фамилия, имя, отчество'
    thanks_amount_cell.value = 'Количество спасибок'
    transactions_amount_cell.value = 'Количество транзакций'
    surname_name_cell.alignment = Alignment(horizontal='center', vertical='center')
    for i in range(len(ADMIN_REPORT_DATA_CELLS)):
        ws[f'{ADMIN_REPORT_DATA_CELLS[i]}3'].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['J2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def get_transactions_data(period: Period) -> Tuple[Dict, Dict, Dict, Dict, QuerySet[UserStat]]:
    stats = UserStat.objects.select_related('user__profile').filter(period=period).only(
        'user_id', 'distr_thanks', 'distr_initial',
        'income_thanks', 'user__profile__first_name',
        'user__profile__middle_name', 'user__profile__surname'
    )
    user_id_list = [stat.user_id for stat in stats]
    transactions = Transaction.objects.filter(
        Q(sender_id__in=user_id_list) |
        Q(recipient_id__in=user_id_list),
        period=period)
    sender_waiting_transactions = (transactions.filter(status__in=['W', 'G'])
                                   .values('sender')
                                   .annotate(waiting=Sum('amount')))
    recipient_waiting_transactions = (transactions.filter(status__in=['W', 'G'])
                                      .values('recipient')
                                      .annotate(waiting=Sum('amount')))
    sender_transactions_amount = transactions.values('sender').annotate(counter=Count('id'))
    recipient_transactions_amount = transactions.values('recipient').annotate(counter=Count('id'))
    sender_waiting_data = {sender['sender']: sender['waiting']
                           for sender in sender_waiting_transactions}
    recipient_waiting_data = {recipient['recipient']: recipient['waiting']
                              for recipient in recipient_waiting_transactions}
    sender_counter_data = {sender['sender']: sender['counter']
                           for sender in sender_transactions_amount}
    recipient_counter_data = {recipient['recipient']: recipient['counter']
                              for recipient in recipient_transactions_amount}
    return recipient_counter_data, recipient_waiting_data, sender_counter_data, sender_waiting_data, stats
